import pandas as pd
from haversine import haversine
from openpyxl import Workbook
from geopy.geocoders import MapBox

WAREHOUSE_LOCATIONS = {
    "SDR Distribution": (43.720850, -79.661280),
    "SDR Distribution - Calgary": (51.158150, -114.000840),
    "Office": (43.6964365, -79.4637599),
    "Second Closet - Toronto": (43.771750, -79.641735),
    "Cambridge": (43.369094, -80.290873),
    "Belleville": (44.162759, -77.383231)
    # Add more warehouses as needed
}


def read_csv_file(file_name):
    return pd.read_csv(file_name, encoding="ISO-8859-1")


def haversine_distance(lat1, lon1, lat2, lon2):
    return haversine((lat1, lon1), (lat2, lon2), unit="km")


def get_coordinates_from_address(address):
    geolocator = MapBox(
        api_key="ENTER API KEY HERE MAPBOX"
    )
    location = geolocator.geocode(address)
    if location:
        print(f"Order with address: {address} has been geocoded sucessfully.")
        return (location.latitude, location.longitude)
    else:
        print(f"ERROR: Could not find coordinates for {address}")
        return None


def update_inventory_level(inventory_df, warehouse_id, sku, change):
    inventory_df.loc[
        (inventory_df["warehouse_id"] == warehouse_id) & (inventory_df["sku"] == sku),
        "inventory_level",
    ] += change


def allocate_orders(inventory_df, orders_df):
    allocations = []
    unallocated_orders = []

    for order_id, order_group in orders_df.groupby("Order ID"):
        sorted_warehouses = sorted(
            [
                (
                    warehouse_id,
                    haversine_distance(
                        order_group.iloc[0]["lat"],
                        order_group.iloc[0]["lon"],
                        coords[0],
                        coords[1],
                    ),
                )
                for warehouse_id, coords in WAREHOUSE_LOCATIONS.items()
            ],
            key=lambda x: x[1],
        )

        # Try to allocate the entire order to a single warehouse first
        allocated = False
        for warehouse_id, distance in sorted_warehouses:
            can_allocate_all = True
            for index, order in order_group.iterrows():
                required_sku = order["SKU"]
                required_quantity = order["Quantity"]

                try:
                    available_quantity = inventory_df.loc[
                        (inventory_df["warehouse_id"] == warehouse_id)
                        & (inventory_df["sku"] == required_sku),
                        "inventory_level",
                    ].values[0]
                except IndexError:
                    available_quantity = 0

                if available_quantity < 0:
                    available_quantity = 0

                if available_quantity < required_quantity:
                    can_allocate_all = False
                    break

            if can_allocate_all:
                for index, order in order_group.iterrows():
                    required_sku = order["SKU"]
                    required_quantity = order["Quantity"]
                    allocations.append(
                        (
                            order["Order ID"],
                            order["SKU"],
                            str(order["Order ID"]) + str(order["SKU"]),
                            warehouse_id,
                            distance,
                            required_quantity,
                            False,
                            order["City"],
                            order["Prov/ State"],
                            order["Postal/ Zip Code"],
                            order["Country Code"],
                        )
                    )
                    update_inventory_level(
                        inventory_df, warehouse_id, required_sku, -required_quantity
                    )
                allocated = True
                break

        # If the entire order can't be allocated to a single warehouse, try to split it
        if not allocated:
            for index, order in order_group.iterrows():
                required_sku = order["SKU"]
                required_quantity = order["Quantity"]

                allocated_quantity = 0

                for warehouse_id, distance in sorted_warehouses:
                    try:
                        available_quantity = inventory_df.loc[
                            (inventory_df["warehouse_id"] == warehouse_id)
                            & (inventory_df["sku"] == required_sku),
                            "inventory_level",
                        ].values[0]
                    except IndexError:
                        available_quantity = 0

                    if available_quantity < 0:
                        available_quantity = 0

                    if available_quantity >= required_quantity - allocated_quantity:
                        allocations.append(
                            (
                                order["Order ID"],
                                order["SKU"],
                                str(order["Order ID"]) + str(order["SKU"]),
                                warehouse_id,
                                distance,
                                required_quantity - allocated_quantity,
                                True,
                                order["City"],
                                order["Prov/ State"],
                                order["Postal/ Zip Code"],
                                order["Country Code"],
                            )
                        )
                        update_inventory_level(
                            inventory_df,
                            warehouse_id,
                            required_sku,
                            -(required_quantity - allocated_quantity),
                        )
                        allocated_quantity = required_quantity
                    else:
                        allocations.append(
                            (
                                order["Order ID"],
                                order["SKU"],
                                str(order["Order ID"]) + str(order["SKU"]),
                                warehouse_id,
                                distance,
                                available_quantity,
                                True,
                                order["City"],
                                order["Prov/ State"],
                                order["Postal/ Zip Code"],
                                order["Country Code"],
                            )
                        )
                        update_inventory_level(
                            inventory_df,
                            warehouse_id,
                            required_sku,
                            -available_quantity,
                        )
                        allocated_quantity += available_quantity

                    if allocated_quantity >= required_quantity:
                        break

                # Add unallocated orders to the list
                if allocated_quantity < required_quantity:
                    unallocated_orders.append(
                        (
                            order["Order ID"],
                            order["SKU"],
                            required_quantity - allocated_quantity,
                        )
                    )

    # Remove allocations with 0 quantity
    allocations = [allocation for allocation in allocations if allocation[5] > 0]

    return allocations, unallocated_orders


def create_excel_report(allocations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocations"

    ws.cell(row=1, column=1, value="Order ID")
    ws.cell(row=1, column=2, value="SKU")
    ws.cell(row=1, column=3, value="VlookupUse")
    ws.cell(row=1, column=4, value="Warehouse ID")
    ws.cell(row=1, column=5, value="Distance (km)")
    ws.cell(row=1, column=6, value="Allocated Quantity")
    ws.cell(row=1, column=7, value="Split Order")
    ws.cell(row=1, column=8, value="City")
    ws.cell(row=1, column=9, value="Prov/ State")
    ws.cell(row=1, column=10, value="Postal/ Zip Code")
    ws.cell(row=1, column=11, value="Country Code")

    for idx, allocation in enumerate(allocations, 2):
        ws.cell(row=idx, column=1, value=allocation[0])
        ws.cell(row=idx, column=2, value=allocation[1])
        ws.cell(row=idx, column=3, value=allocation[2])
        ws.cell(row=idx, column=4, value=allocation[3])
        ws.cell(row=idx, column=5, value=allocation[4])
        ws.cell(row=idx, column=6, value=allocation[5])
        ws.cell(row=idx, column=7, value=allocation[6])
        ws.cell(row=idx, column=8, value=allocation[7])
        ws.cell(row=idx, column=9, value=allocation[8])
        ws.cell(row=idx, column=10, value=allocation[9])
        ws.cell(row=idx, column=11, value=allocation[10])

    wb.save("allocation_report.xlsx")


import tkinter as tk
from tkinter import filedialog, messagebox


class WarehouseAllocatorGUI(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Warehouse Allocator")

        self.inventory_file = None
        self.orders_file = None

        self.create_widgets()

    def create_widgets(self):
        self.inventory_label = tk.Label(self, text="Inventory CSV File:")
        self.inventory_label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)

        self.inventory_button = tk.Button(
            self, text="Choose File", command=self.load_inventory
        )
        self.inventory_button.grid(row=0, column=1, padx=10, pady=10)

        self.orders_label = tk.Label(self, text="Orders CSV File:")
        self.orders_label.grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)

        self.orders_button = tk.Button(
            self, text="Choose File", command=self.load_orders
        )
        self.orders_button.grid(row=1, column=1, padx=10, pady=10)

        self.start_button = tk.Button(
            self, text="Start Allocation", command=self.start_allocation
        )
        self.start_button.grid(row=2, column=0, padx=10, pady=10, columnspan=2)

        self.quit_button = tk.Button(self, text="Quit", command=self.quit_program)
        self.quit_button.grid(row=3, column=0, padx=10, pady=10, columnspan=2)

    def load_inventory(self):
        self.inventory_file = filedialog.askopenfilename()
        self.inventory_label.config(text=f"Inventory CSV File: {self.inventory_file}")

    def load_orders(self):
        self.orders_file = filedialog.askopenfilename()
        self.orders_label.config(text=f"Orders CSV File: {self.orders_file}")

    def start_allocation(self):
        print("Running...")
        if self.inventory_file and self.orders_file:
            inventory_df = read_csv_file(self.inventory_file)
            orders_df = read_csv_file(self.orders_file)

            # Add your code for processing data (like adding lat and lon columns) here
            # Add latitude and longitude columns to the orders DataFrame
            orders_df["lat"] = 0.0
            orders_df["lon"] = 0.0
            for idx, order in orders_df.iterrows():
                address = (
                    f"{order['City']}, {order['Prov/ State']}, {order['Country Code']}"
                )
                coordinates = get_coordinates_from_address(address)
                if coordinates:
                    orders_df.at[idx, "lat"] = coordinates[0]
                    orders_df.at[idx, "lon"] = coordinates[1]

            allocations, unallocated_orders = allocate_orders(inventory_df, orders_df)
            create_excel_report(allocations)
            messagebox.showinfo(
                "Success", "Allocation report generated: allocation_report.xlsx"
            )

            if unallocated_orders:
                unallocated_orders_str = "\n".join(
                    [
                        f"Order ID: {order[0]}, SKU: {order[1]}, Quantity: {order[2]}"
                        for order in unallocated_orders
                    ]
                )
                print(f"Unallocated Orders:\n\n{unallocated_orders_str}")
            else:
                messagebox.showinfo(
                    "All orders allocated", "All orders have been allocated."
                )
        else:
            messagebox.showerror(
                "Error", "Please select both inventory and orders CSV files."
            )

    def quit_program(self):
        self.destroy()


def main():
    app = WarehouseAllocatorGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
